import streamlit as st
import pandas as pd
import datetime
import time
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- 設定 ---
st.set_page_config(
    page_title="勤怠管理アプリ",
    page_icon="⏰",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 定数（Excelの仕様に合わせて調整）
WORK_HOURS_PER_DAY = 7.5  # 所定労働時間
NIGHT_START_HOUR = 22     # 深夜開始
NIGHT_END_HOUR = 5        # 深夜終了

# --- データベース接続 (Firestore) ---
if not firebase_admin._apps:
    if "firebase" in st.secrets:
        cred_info = dict(st.secrets["firebase"])
        cred = credentials.Certificate(cred_info)
        firebase_admin.initialize_app(cred)
    else:
        st.error("【重要】Firebase認証情報が設定されていません。Streamlit Secretsを設定してください。")
        st.stop()

db = firestore.client()

# --- ユーティリティ ---
import hashlib
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def get_current_time_str():
    return datetime.datetime.now().strftime("%H:%M")

def get_today_str():
    return datetime.date.today().strftime("%Y-%m-%d")

# --- 時間計算ロジック（深夜・残業・日跨ぎ対応） ---
def calculate_work_stats(clock_in, clock_out, break_start=None, break_end=None):
    """
    始業・終業・休憩から、実働、残業、深夜時間を計算する高精度ロジック
    戻り値: (実働時間h, 残業時間h, 深夜時間h) すべてfloat
    """
    if not clock_in or not clock_out:
        return 0.0, 0.0, 0.0
    
    # datetimeオブジェクトへ変換 (日付は今日と仮定、日跨ぎ対応)
    fmt = "%H:%M"
    base_date = datetime.datetime.today().date()
    try:
        t_in = datetime.datetime.strptime(clock_in, fmt).replace(year=base_date.year, month=base_date.month, day=base_date.day)
        t_out = datetime.datetime.strptime(clock_out, fmt).replace(year=base_date.year, month=base_date.month, day=base_date.day)
    except ValueError:
        return 0.0, 0.0, 0.0

    if t_out < t_in:
        t_out += datetime.timedelta(days=1) # 日跨ぎ
        
    # 休憩時間の計算
    break_hours = 0.0
    if break_start and break_end:
        try:
            b_in = datetime.datetime.strptime(break_start, fmt).replace(year=base_date.year, month=base_date.month, day=base_date.day)
            b_out = datetime.datetime.strptime(break_end, fmt).replace(year=base_date.year, month=base_date.month, day=base_date.day)
            if b_out < b_in:
                b_out += datetime.timedelta(days=1)
            
            # 休憩が勤務時間内かチェック（簡易）
            break_hours = (b_out - b_in).total_seconds() / 3600
        except ValueError:
            pass
    
    # 総拘束時間
    total_duration = (t_out - t_in).total_seconds() / 3600
    
    # 実労働時間
    net_work_hours = max(0.0, total_duration - break_hours)
    
    # 残業時間 (7.5時間超)
    overtime_hours = max(0.0, net_work_hours - WORK_HOURS_PER_DAY)
    
    # 深夜時間 (22:00 - 05:00) の判定
    # 1分ごとに判定（精度重視）
    night_minutes = 0
    current = t_in
    while current < t_out:
        # ※本来は休憩時間中を除外する処理が必要ですが、今回は簡易的に全期間で判定します
        h = current.hour
        if h >= NIGHT_START_HOUR or h < NIGHT_END_HOUR:
            night_minutes += 1
        current += datetime.timedelta(minutes=1)
    
    night_hours = night_minutes / 60.0
    
    return net_work_hours, overtime_hours, night_hours

def format_hour(val):
    """Excel表示用: 7.5 -> 7:30 形式の文字列に変換"""
    if val is None or val == 0:
        return ""
    hours = int(val)
    minutes = int((val - hours) * 60)
    return f"{hours:02}:{minutes:02}"

# --- データベース操作関数 ---
def get_employee(name):
    docs = db.collection('employees').where('name', '==', name).stream()
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        return data
    return None

def get_employee_by_id(doc_id):
    doc = db.collection('employees').document(doc_id).get()
    if doc.exists:
        data = doc.to_dict()
        data['id'] = doc.id
        return data
    return None

def get_all_employees():
    docs = db.collection('employees').stream()
    employees = []
    for doc in docs:
        data = doc.to_dict()
        data['id'] = doc.id
        employees.append(data)
    return employees

def get_admin(username):
    docs = db.collection('admins').where('username', '==', username).stream()
    for doc in docs:
        return doc.to_dict()
    return None

def get_attendance_range(employee_id, start_date, end_date):
    """指定期間の勤怠データを取得"""
    # 日付文字列でクエリ
    docs = db.collection('attendance')\
             .where('employee_id', '==', employee_id)\
             .where('date', '>=', str(start_date))\
             .where('date', '<=', str(end_date))\
             .stream()
    
    data_list = []
    for doc in docs:
        d = doc.to_dict()
        d['doc_id'] = doc.id
        data_list.append(d)
    
    # 日付順にソート
    data_list.sort(key=lambda x: x['date'])
    return data_list

def get_attendance_today(employee_id, date_str):
    docs = db.collection('attendance')\
             .where('employee_id', '==', employee_id)\
             .where('date', '==', date_str)\
             .stream()
    for doc in docs:
        data = doc.to_dict()
        data['doc_id'] = doc.id
        return data
    return None

# --- Excel生成エンジン (openpyxl) ---
def generate_monthly_report_excel(employee_data, year, month, records):
    """
    提供されたExcelフォーマット（勤務月報査定表）をopenpyxlで再現描画
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月_{employee_data['name']}"
    
    # --- スタイル定義 ---
    font_title = Font(name='ＭＳ ゴシック', size=16, bold=True)
    font_header = Font(name='ＭＳ ゴシック', size=11, bold=True)
    font_body = Font(name='ＭＳ ゴシック', size=10)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_header = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # --- ヘッダー部 ---
    # タイトル
    ws.merge_cells('A1:X1')
    ws['A1'] = "【　勤　務　月　報　査　定　表　】"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    # 所属・氏名など (3行目〜)
    ws['A3'] = f"{year}"
    ws['C3'] = "年"
    ws['D3'] = f"{month}"
    ws['F3'] = "月"
    
    ws['M3'] = "氏名"
    ws['O3'] = employee_data['name']
    ws['M4'] = "所属"
    ws['O4'] = "ＣＨＥｚＬｅＰａｉｎ山形店" 
    
    # 装飾
    for cell in ['M3', 'O3', 'M4', 'O4']:
        ws[cell].border = border_thin
        ws[cell].font = font_body

    # --- テーブルヘッダー (2段組み・結合) ---
    headers_def = [
        ('A7:A8', '日付'), ('B7:B8', '曜日'), 
        ('C7:E7', '就業時間'), 
        ('F7:H7', '勤務'), 
        ('I7:K7', '超過勤務'), 
        ('L7:N7', '法定内休日'), 
        ('O7:Q7', '法定外休日'),
        ('R7:R8', '記事'), ('S7:S8', '備考')
    ]
    
    # 大項目描画
    for rng, val in headers_def:
        ws.merge_cells(rng)
        top_left = rng.split(':')[0]
        cell = ws[top_left]
        cell.value = val
        cell.alignment = align_center
        cell.border = border_thin
        cell.fill = fill_header
        cell.font = font_header

    # 中項目 (8行目)
    sub_headers = {
        'C8': '始業', 'D8': '終業', 'E8': '休憩',
        'F8': '実働', 'G8': '移動', 'H8': '時間内',
        'I8': '残業', 'J8': '深夜', 'K8': '時間内', 
        # Excelの構造に合わせて拡張
    }

    for cell_addr, val in sub_headers.items():
        cell = ws[cell_addr]
        cell.value = val
        cell.alignment = align_center
        cell.border = border_thin
        cell.font = font_body

    # --- データ行の描画 ---
    import calendar
    try:
        last_day = calendar.monthrange(year, month)[1]
    except:
        last_day = 30
    
    # 勤怠データを辞書化 (日付文字列キー)
    att_map = {r['date']: r for r in records}
    
    row_idx = 9
    total_net = 0.0
    total_over = 0.0
    total_night = 0.0
    
    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    
    for day in range(1, last_day + 1):
        date_obj = datetime.date(year, month, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        
        # A列: 日付, B列: 曜日
        ws[f'A{row_idx}'] = day
        ws[f'B{row_idx}'] = weekdays_jp[date_obj.weekday()]
        
        # 罫線適用
        for col in range(1, 20): # A to S
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_thin
            cell.alignment = align_center
            cell.font = font_body

        # 実績埋め込み
        if date_str in att_map:
            d = att_map[date_str]
            
            # 時間計算
            net, over, night = calculate_work_stats(
                d.get('clock_in'), d.get('clock_out'), 
                d.get('break_start'), d.get('break_end')
            )
            
            ws[f'C{row_idx}'] = d.get('clock_in', '')
            ws[f'D{row_idx}'] = d.get('clock_out', '')
            
            # 休憩表示 (簡易)
            if d.get('break_start'):
                ws[f'E{row_idx}'] = f"{d.get('break_start')}~"
            
            ws[f'F{row_idx}'] = format_hour(net)
            ws[f'I{row_idx}'] = format_hour(over) # 残業列
            ws[f'J{row_idx}'] = format_hour(night) # 深夜列
            
            total_net += net
            total_over += over
            total_night += night
            
        row_idx += 1

    # --- 合計行 ---
    ws[f'A{row_idx}'] = "合　計"
    ws.merge_cells(f'A{row_idx}:B{row_idx}')
    cell_sum = ws[f'A{row_idx}']
    cell_sum.alignment = align_center
    cell_sum.border = border_thin
    cell_sum.font = font_header
    
    # 合計値描画
    ws[f'F{row_idx}'] = format_hour(total_net)
    ws[f'I{row_idx}'] = format_hour(total_over)
    ws[f'J{row_idx}'] = format_hour(total_night)
    
    for col in range(3, 20):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = border_thin
        cell.font = font_body

    # 列幅調整 (簡易)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 5
    for c in ['C','D','E','F','G','H','I','J','K']:
         ws.column_dimensions[c].width = 10
    
    return wb

# --- UIスタイル ---
def style_setup():
    st.markdown("""
    <style>
        /* Google Fonts: M PLUS Rounded 1c (丸ゴシック) */
        @import url('https://fonts.googleapis.com/css2?family=M+PLUS+Rounded+1c:wght@400;700&display=swap');

        html, body, [class*="css"] {
            font-family: 'M PLUS Rounded 1c', sans-serif;
        }

        /* タイトル */
        h1 {
            color: #FF8BA7; 
            text-shadow: 2px 2px 0px #FFF0F5;
        }
        
        /* ボタン */
        .stButton>button {
            width: 100%;
            border-radius: 50px;
            border: none;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            transition: all 0.2s;
            font-weight: bold;
        }
        .stButton>button:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 8px rgba(0,0,0,0.15);
        }
        
        /* カラースキーム */
        /* 出勤系 */
        div[data-testid="column"]:nth-of-type(1) .stButton>button {
            background-color: #A0E7E5; color: #333;
        }
        /* 退勤系 */
        div[data-testid="column"]:nth-of-type(2) .stButton>button {
            background-color: #FFAEBC; color: #333;
        }
        /* 休憩系 */
        div[data-testid="column"]:nth-of-type(3) .stButton>button {
            background-color: #FBE7C6; color: #333;
        }
        /* 再開系 */
        div[data-testid="column"]:nth-of-type(4) .stButton>button {
            background-color: #B4F8C8; color: #333;
        }
    </style>
    """, unsafe_allow_html=True)

# --- 画面: 認証 ---
def login_screen():
    st.title("勤怠管理アプリ 🍩")
    
    admins = db.collection('admins').limit(1).stream()
    if not list(admins):
        st.warning("管理者が登録されていません。")
        if st.button("初期管理者作成"):
            hashed = hash_password("password")
            db.collection('admins').add({
                "username": "admin",
                "password": hashed
            })
            st.success("作成しました。")
            time.sleep(2)
            st.rerun()

    tab1, tab2 = st.tabs(["🐣 スタッフ", "🔧 管理者"])
    
    with tab1:
        st.header("さあ、はじめましょう！")
        employees = get_all_employees()
        if not employees:
            st.info("スタッフが登録されていません。")
        else:
            emp_names = [e['name'] for e in employees]
            selected_name = st.selectbox("お名前を選んでください", emp_names)
            pin = st.text_input("暗証番号 (4桁)", type="password", key="staff_pin", max_chars=4)
            
            c1, c2, c3 = st.columns([1, 2, 1])
            with c2:
                if st.button("スタート ▶︎", key="staff_login_btn"):
                    emp_data = get_employee(selected_name)
                    if emp_data and emp_data.get('pin') == pin:
                        st.session_state['logged_in'] = True
                        st.session_state['user_role'] = 'staff'
                        st.session_state['user_id'] = emp_data['id']
                        st.session_state['user_name'] = selected_name
                        st.rerun()
                    else:
                        st.error("暗証番号が違います🥺")

    with tab2:
        st.header("管理者ログイン")
        admin_user = st.text_input("管理者ID")
        admin_pass = st.text_input("パスワード", type="password")
        
        c1, c2, c3 = st.columns([1, 2, 1])
        with c2:
            if st.button("ログイン", key="admin_login_btn"):
                admin_data = get_admin(admin_user)
                if admin_data and admin_data['password'] == hash_password(admin_pass):
                    st.session_state['logged_in'] = True
                    st.session_state['user_role'] = 'admin'
                    st.session_state['user_name'] = admin_user
                    st.rerun()
                else:
                    st.error("IDまたはパスワードが違います")

# --- 画面: スタッフ機能 ---
def staff_dashboard():
    st.title(f"お疲れ様です、{st.session_state['user_name']}さん ✨")
    
    today = get_today_str()
    record = get_attendance_today(st.session_state['user_id'], today)
    
    # recordがNoneの場合の対策
    clock_in = record.get('clock_in') if record else None
    clock_out = record.get('clock_out') if record else None
    break_start = record.get('break_start') if record else None
    break_end = record.get('break_end') if record else None
    doc_id = record.get('doc_id') if record else None

    # ステータス表示
    st.markdown("### 📅 今日のステータス")
    c1, c2 = st.columns(2)
    c1.metric("出勤時刻", clock_in if clock_in else "--:--")
    c2.metric("退勤時刻", clock_out if clock_out else "--:--")

    st.write("") 

    photo = st.camera_input("認証用写真撮影", label_visibility="collapsed")
    photo_b64 = None
    if photo:
        photo_b64 = base64.b64encode(photo.getvalue()).decode()

    st.write("")

    col1, col2 = st.columns(2)
    col3, col4 = st.columns(2)
    
    with col1:
        if st.button("☀️ 出勤"):
            if not photo_b64:
                st.warning("写真を撮影してください📸")
            elif clock_in:
                st.warning("すでに出勤しています")
            else:
                db.collection('attendance').add({
                    'employee_id': st.session_state['user_id'],
                    'date': today,
                    'clock_in': get_current_time_str(),
                    'photo': photo_b64,
                    'created_at': firestore.SERVER_TIMESTAMP
                })
                st.success("おはようございます！今日も頑張りましょう！🌈")
                time.sleep(2)
                st.rerun()

    with col2:
        if st.button("🌙 退勤"):
            if not clock_in:
                st.warning("まだ出勤していません")
            elif clock_out:
                st.warning("すでに退勤しています")
            else:
                db.collection('attendance').document(doc_id).update({
                    'clock_out': get_current_time_str()
                })
                st.success("お疲れ様でした！ゆっくり休んでください🍵")
                time.sleep(2)
                st.rerun()
    
    with col3:
        if st.button("☕️ 休憩"):
            if doc_id and not break_start:
                db.collection('attendance').document(doc_id).update({
                    'break_start': get_current_time_str()
                })
                st.rerun()
            else:
                st.warning("操作できません")

    with col4:
        if st.button("💪 再開"):
            if doc_id and break_start and not break_end:
                db.collection('attendance').document(doc_id).update({
                    'break_end': get_current_time_str()
                })
                st.rerun()
            else:
                st.warning("操作できません")

    st.divider()

    with st.expander("💰 今月の概算給与"):
        emp = get_employee_by_id(st.session_state['user_id'])
        current_month = datetime.datetime.now().strftime("%Y-%m")
        start_m = current_month + "-01"
        end_m = current_month + "-31"
        
        logs = db.collection('attendance')\
                 .where('employee_id', '==', st.session_state['user_id'])\
                 .where('date', '>=', start_m)\
                 .where('date', '<=', end_m)\
                 .stream()
        
        work_hours = 0.0
        for log in logs:
            d = log.to_dict()
            net, _, _ = calculate_work_stats(d.get('clock_in'), d.get('clock_out'), d.get('break_start'), d.get('break_end'))
            work_hours += net
        
        est_pay = 0
        if emp and emp.get('salary_type') == '月給':
            est_pay = emp.get('salary', 0)
        elif emp:
            est_pay = int(work_hours * emp.get('salary', 0))
            
        if st.checkbox("金額を表示する"):
            st.metric("概算給与", f"{est_pay:,} 円")
        else:
            st.metric("概算給与", "***** 円")

# --- 画面: 管理者機能 ---
def admin_dashboard():
    st.title("管理者ダッシュボード 🛠️")
    menu = st.sidebar.radio("メニュー", ["👥 スタッフ管理", "👤 個人実績・出力", "✏️ 勤怠修正", "📊 全体集計", "⚙️ システム設定"])

    if menu == "👥 スタッフ管理":
        st.subheader("スタッフ登録")
        with st.form("add_emp"):
            c1, c2 = st.columns(2)
            name = c1.text_input("氏名")
            birth = c2.date_input("生年月日", min_value=datetime.date(1960, 1, 1))
            c3, c4 = st.columns(2)
            e_type = c3.selectbox("雇用形態", ["社員", "AP"])
            s_type = c4.selectbox("給与形態", ["月給", "時給"])
            c5, c6 = st.columns(2)
            salary = c5.number_input("給与額", min_value=0)
            trans = c6.number_input("交通費", min_value=0)
            pin = st.text_input("暗証番号 (4桁)", max_chars=4)
            
            if st.form_submit_button("登録"):
                db.collection('employees').add({
                    'name': name,
                    'birth_date': str(birth),
                    'employee_type': e_type,
                    'salary_type': s_type,
                    'salary': salary,
                    'transportation': trans,
                    'pin': pin,
                    'created_at': firestore.SERVER_TIMESTAMP
                })
                st.success("登録しました")
                time.sleep(1)
                st.rerun()

        st.subheader("登録済みスタッフ")
        emps = get_all_employees()
        if emps:
            df = pd.DataFrame(emps)
            st.dataframe(df[['name', 'employee_type', 'salary_type', 'id']])
            
            # 従業員マスタ出力
            output_emp = BytesIO()
            with pd.ExcelWriter(output_emp, engine='openpyxl') as writer:
                valid_cols = [c for c in ['id', 'name', 'birth_date', 'employee_type', 'salary_type', 'salary', 'transportation', 'pin'] if c in df.columns]
                df[valid_cols].to_excel(writer, sheet_name='従業員マスタ', index=False)
            output_emp.seek(0)
            st.download_button("従業員マスタ Excel出力", data=output_emp, file_name="employee_master.xlsx")

            del_id = st.selectbox("削除対象ID", [e['id'] for e in emps])
            if st.button("選択したスタッフを削除"):
                db.collection('employees').document(del_id).delete()
                st.warning("削除しました")
                time.sleep(1)
                st.rerun()

    elif menu == "👤 個人実績・出力":
        st.subheader("個人別勤怠レポート & Excel出力")
        st.info("スタッフを選択して、勤務表（Excel原本形式）をダウンロードできます。")
        
        employees = get_all_employees()
        if employees:
            c1, c2 = st.columns(2)
            emp_map = {e['name']: e for e in employees}
            sel_name = c1.selectbox("スタッフ", list(emp_map.keys()))
            target_emp = emp_map[sel_name]
            
            today = datetime.date.today()
            sel_month = c2.date_input("対象年月", value=today)
            
            # データ取得
            import calendar
            start_date = sel_month.replace(day=1)
            last_day = calendar.monthrange(start_date.year, start_date.month)[1]
            end_date = start_date.replace(day=last_day)
            
            records = get_attendance_range(target_emp['id'], start_date, end_date)
            
            # 簡易プレビュー
            st.markdown(f"**{sel_name}** さんの **{start_date.year}年{start_date.month}月** の実績")
            
            prev_data = []
            total_net = 0.0
            
            for d in records:
                net, over, night = calculate_work_stats(d.get('clock_in'), d.get('clock_out'), d.get('break_start'), d.get('break_end'))
                prev_data.append({
                    "日付": d['date'],
                    "出勤": d.get('clock_in'),
                    "退勤": d.get('clock_out'),
                    "実働": format_hour(net),
                    "残業": format_hour(over),
                    "深夜": format_hour(night)
                })
                total_net += net
            
            if prev_data:
                st.dataframe(pd.DataFrame(prev_data))
                st.metric("合計実労働時間", format_hour(total_net))
            else:
                st.warning("データがありません")
            
            # Excel出力ボタン
            wb = generate_monthly_report_excel(target_emp, start_date.year, start_date.month, records)
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            
            st.download_button(
                label="📥 勤務表をExcelでダウンロード",
                data=out,
                file_name=f"勤怠管理表_{sel_name}_{start_date.month}月.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    elif menu == "✏️ 勤怠修正":
        st.subheader("勤怠データの修正・追加")
        emps = get_all_employees()
        if emps:
            c1, c2 = st.columns(2)
            selected_emp_id = c1.selectbox("スタッフ選択", [e['id'] for e in emps], format_func=lambda x: next(e['name'] for e in emps if e['id'] == x))
            selected_date = c2.date_input("日付選択", value=datetime.date.today())
            date_str = str(selected_date)

            record = get_attendance_today(selected_emp_id, date_str)
            
            def_in = datetime.time(9, 0)
            def_out = datetime.time(18, 0)
            def_b_start = None
            def_b_end = None
            doc_id = None
            
            if record:
                st.write("📝 データあり。修正モード")
                doc_id = record['doc_id']
                if record.get('clock_in'):
                    def_in = datetime.datetime.strptime(record['clock_in'], "%H:%M").time()
                if record.get('clock_out'):
                    def_out = datetime.datetime.strptime(record['clock_out'], "%H:%M").time()
            else:
                st.info("新規作成モード")

            with st.form("edit_attendance"):
                tc1, tc2 = st.columns(2)
                new_in = tc1.time_input("出勤", value=def_in)
                new_out = tc2.time_input("退勤", value=def_out)
                
                if st.form_submit_button("保存"):
                    data = {
                        'clock_in': new_in.strftime("%H:%M"),
                        'clock_out': new_out.strftime("%H:%M"),
                        'date': date_str,
                        'employee_id': selected_emp_id
                    }
                    if doc_id:
                        db.collection('attendance').document(doc_id).update(data)
                        st.success("更新しました")
                    else:
                        data['created_at'] = firestore.SERVER_TIMESTAMP
                        db.collection('attendance').add(data)
                        st.success("作成しました")
                    time.sleep(1)
                    st.rerun()

    elif menu == "📊 全体集計":
        st.subheader("月間データ出力（旧仕様）")
        d1, d2 = st.columns(2)
        start_d = d1.date_input("開始", value=datetime.date.today().replace(day=1))
        end_d = d2.date_input("終了", value=datetime.date.today())
        
        if st.button("一覧ダウンロード"):
            all_logs = db.collection('attendance').stream()
            data_list = []
            emp_map = {e['id']: e for e in get_all_employees()}
            for doc in all_logs:
                d = doc.to_dict()
                log_date = datetime.datetime.strptime(d['date'], "%Y-%m-%d").date()
                if start_d <= log_date <= end_d:
                    emp = emp_map.get(d['employee_id'])
                    if emp:
                        ymd = d['date'].split('-')
                        data_list.append({
                            '名前': emp['name'],
                            '年': int(ymd[0]),
                            '月': int(ymd[1]),
                            '日': int(ymd[2]),
                            '出勤': d.get('clock_in'),
                            '退勤': d.get('clock_out'),
                            '給与形態': emp['salary_type']
                        })
            if data_list:
                df_res = pd.DataFrame(data_list)
                st.dataframe(df_res)
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_res.to_excel(writer, sheet_name='勤怠一覧', index=False)
                output.seek(0)
                st.download_button("Excelダウンロード", data=output, file_name="attendance_list.xlsx")
            else:
                st.warning("データなし")

    elif menu == "⚙️ システム設定":
        st.info("管理者パスワード変更など")
        new_p = st.text_input("新パスワード", type="password")
        if st.button("変更"):
            docs = db.collection('admins').where('username', '==', 'admin').stream()
            for doc in docs:
                db.collection('admins').document(doc.id).update({'password': hash_password(new_p)})
            st.success("変更しました")

# --- メイン実行 ---
def main():
    style_setup()
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False

    if st.session_state['logged_in']:
        with st.sidebar:
            st.write(f"User: {st.session_state.get('user_name')}")
            if st.button("ログアウト"):
                st.session_state.clear()
                st.rerun()

    if not st.session_state['logged_in']:
        login_screen()
    else:
        if st.session_state['user_role'] == 'staff':
            staff_dashboard()
        else:
            admin_dashboard()

if __name__ == "__main__":
    main()
